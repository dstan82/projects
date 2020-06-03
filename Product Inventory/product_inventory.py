from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter


class Product():
    def __init__(self, id, name, price, quantity):
        self.id = id
        self.name = name
        self.price = price
        self.quantity = quantity

    def add_quantity(self, quantity):
        self.quantity += quantity


class Inventory():
    def __init__(self):
        self.all_products = []

    def add_product(self, product):
        for elem in self.all_products:
            if product.id == elem.id:
                elem.add_quantity(product.quantity)
                return 'Product already exists, adding quantity'
        self.all_products.append(product)
        return 'Product not found, adding to inventory list'

    def product_value(self, product):
        return product.price * product.quantity


inventory = Inventory()  # Creating an inventory object

wb = load_workbook('inventory.xlsx')  # open the product xlsx file
sheet = wb.active

table_headers = sheet['A1':get_column_letter(sheet.max_column)+'1']  # reading only the table headers
all_products = sheet['A2':get_column_letter(sheet.max_column)+str(sheet.max_row)]  # reading all the products

for a, b, c, d in all_products:
    product = Product(a.value, b.value, c.value, d.value)  # instantiating all the products
    inventory.add_product(product)  # Adding the newly created product in inventory

workbook = Workbook()  # Creating a new xlsx file 'Workbook'
sheet = workbook.active  # Selecting active sheet
sheet.title = 'Inventory Report'  # Rename the active sheet

sheet['A1'] = 'Inventory Report, containing unique IDs and total values for each ID'

#  Appending the table headers
headers_row = []
for header_item in table_headers[0]:
    headers_row.append(header_item.value)
sheet.append(headers_row)

sheet['A3'] = '-'*100

#  Appending all the products
for product in inventory.all_products:
    sheet.append([product.id, product.name, product.price, product.quantity, 'Total value:', inventory.product_value(product)])

#  Calculating and appending total inventory value (value of ALL products)
total_value = 0
for product in inventory.all_products:
    total_value += inventory.product_value(product)
sheet.append(['_'*68])
sheet.append(['Total goods value:', '', '', '', '', total_value])

#  Saving file
filename = "report.xlsx"
workbook.save(filename)
